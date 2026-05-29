# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path

# Paths
TEMPLATE_PATH = Path(r"F:/Sem 4/DAP391/AI_AuditLog_Template_DAP391m.xlsx")

print("Loading workbook...")
wb = openpyxl.load_workbook(TEMPLATE_PATH)

# ==============================================================================
# SHEET 1: 1. Metadata & Summary
# ==============================================================================
print("Populating Sheet 1: Metadata & Summary...")
ws1 = wb['1. Metadata & Summary']

# Student Information
ws1['C4'] = "Nguyễn Trọng Nhân"
ws1['C5'] = "[Mã SV của Nhân]" # Placeholder for student to replace
ws1['C6'] = "DAP391m"
ws1['C7'] = "Multi-Channel Attribution Modeling Project"

# AI Usage Summary
ws1['C10'] = 148
ws1['C11'] = 18
ws1['C13'] = 3

# AI Tools Used Table (Rows 17 to 20)
# Column A: AI Tool, Column B: Purpose, Column C: Frequency, Column D: Main Value
tools_data = [
    ("ChatGPT", "Nghiên cứu tổng quan lý thuyết, brainstorm phương pháp và công thức hồi quy.", "High", "Hỗ trợ xây dựng khung nghiên cứu"),
    ("Claude", "Hỗ trợ debug logic dữ liệu chuỗi hành trình phức tạp và lập luận trong Human Reflection.", "High", "Đóng góp giá trị phản biện học thuật"),
    ("GitHub Copilot", "Tự động hóa sinh mã nguồn boilerplate trong các cell Jupyter Notebook.", "High", "Tăng tốc độ viết mã"),
    ("Gemini", "Rà soát chéo các lỗi logic toán học và xác minh độ chuẩn xác của tài liệu tham khảo.", "Medium", "Kiểm chứng thông tin độc lập")
]

for idx, (tool, purpose, freq, val) in enumerate(tools_data, start=17):
    ws1[f'A{idx}'] = tool
    ws1[f'B{idx}'] = purpose
    ws1[f'C{idx}'] = freq
    ws1[f'D{idx}'] = val

# ==============================================================================
# SHEET 2: 2. Detailed Audit Log
# ==============================================================================
print("Populating Sheet 2: Detailed Audit Log...")
ws2 = wb['2. Detailed Audit Log']

# Remove existing placeholder rows (starting from row 4 onwards)
# We will just write starting from Row 4.
entries = [
    {
        "num": "1",
        "type": "DECISION",
        "stage": "Business & Problem Understanding",
        "context": "Xác định xem có nên đưa mô hình Markov Chain làm cấu phần cốt lõi của dự án khi timeline coding chỉ còn 3 ngày trước hạn báo cáo.",
        "prompt": "Nhóm em đang làm project DAP391m đề tài Multi-channel Attribution. Bọn em định chạy xài Markov Chain làm mô hình so sánh cốt lõi cùng Heuristic, nhưng timeline chỉ còn 3 ngày cho coding. Có nên để Markov làm core deliverable không hay chuyển làm stretch goal?",
        "response": "AI khuyên nên giữ Markov Chain làm core deliverable vì đây là mô hình nâng cao thể hiện tính học thuật mạnh mẽ, khuyên dùng thư viện NetworkX để vẽ đồ thị transition matrix.",
        "reflection": "Critical Thinking: AI khuyên dùng Markov làm core nhưng bỏ qua rủi ro thời gian và việc debug thư viện NetworkX khi dữ liệu có nhiều chuỗi hành trình phức tạp.\nContextualization: Dữ liệu thực tế nhóm chỉ có 3 ngày trước deadline 28/05. Khang chưa hoàn thiện xong data_cleaning.py, nếu nhồi thêm Markov làm core sẽ dẫn đến trễ toàn bộ pipeline.\nCreative Synthesis: Đưa Markov Chain về dạng Stretch Goal và chỉ thực thi ở notebook phụ 03_attribution_markov.ipynb sau khi đã chốt xong 3 RQ cốt lõi.\nDecision Ownership: Quyết định dứt khoát đưa Markov thành Stretch Goal, tập trung tối đa nguồn lực làm tốt 3 Heuristic models (FT/LT/Linear) và Logistic regression trước để đảm bảo tính an toàn cho timeline.",
        "evidence": "notebooks/README.md phần \"Scope and Roadmap\" mô tả Markov Chain là stretch goal."
    },
    {
        "num": "2",
        "type": "DECISION",
        "stage": "Business & Problem Understanding",
        "context": "Định hướng bài nghiên cứu khoa học khi phát hiện tín hiệu độc lập giữa Channel và Conversion quá mạnh (phép thử Chi-square p > 0.85).",
        "prompt": "Dataset Attribution của em có kết quả Chi-square giữa Channel và Conversion là p=0.87. AI khuyên em nên dùng các mô hình dự báo nâng cao để forecast doanh thu marketing. Em nên làm thế nào?",
        "response": "AI đề xuất tiếp tục thử nghiệm Random Forest hoặc Neural Network để dự báo doanh số thực tế vì nó có thể bắt được các tương tác phi tuyến mà Chi-square bỏ qua.",
        "reflection": "Critical Thinking: AI đề xuất quá lạc quan về sức mạnh mô hình phi tuyến. Khi p-value của Chi-square quá cao (>0.85), các kênh độc lập cơ bản với biến conversion ở mức touchpoint, việc cố gắng dự báo doanh thu tuyệt đối từ dữ liệu này là vô nghĩa và thiếu cơ sở khoa học (overfitting cực nặng).\nContextualization: Báo cáo viết theo chuẩn nghiên cứu khoa học, cần đảm bảo tính logic và trung thực về học thuật để bảo vệ trước hội đồng giảng viên.\nCreative Synthesis: Chuyển trọng tâm của bài paper từ \"dự báo doanh số tuyệt đối\" sang \"phương pháp so sánh tính phù hợp giữa các heuristic attribution models\" và bàn luận sâu về ý nghĩa của việc signal yếu đối với thực tiễn phân bổ ngân sách.\nDecision Ownership: Quyết định thay đổi định hướng nghiên cứu, lấy signal yếu làm điểm nhấn học thuật (finding quan trọng) trong mục Discussion & Limitations thay vì cố tạo ra một forecast model giả tạo.",
        "evidence": "DAP391_Analysis_Report.html phần 3.1 và 3.3 liệt kê kết quả Chi-square test độc lập."
    },
    {
        "num": "3",
        "type": "DECISION",
        "stage": "Business & Problem Understanding",
        "context": "Lựa chọn cấu trúc lưu trữ và tổ chức mã nguồn để tối ưu hóa quy trình làm việc nhóm và tính Reproducibility của 3 câu hỏi nghiên cứu (RQ).",
        "prompt": "Em nên code chung toàn bộ phân tích attribution, logistic regression và simulation vào một file Jupyter Notebook hay chia nhỏ ra?",
        "response": "AI khuyên nên gộp chung tất cả vào một file Jupyter notebook duy nhất để dễ quản lý luồng code, tránh mất thời gian load data nhiều lần và tiện xuất ra báo cáo PDF một lượt.",
        "reflection": "Critical Thinking: AI khuyên gộp chung để tiện nhưng bỏ qua việc một notebook quá dài sẽ rất khó debug, dễ gây xung đột biến (ví dụ trùng tên biến df hay model) và cản trở việc làm việc nhóm song song.\nContextualization: Nhân phụ trách coding chính nhưng cần Khang cung cấp các file dữ liệu làm sạch. Nếu viết chung một notebook, việc tái tạo kết quả cho từng RQ riêng biệt khi có bản cập nhật data mới từ Khang sẽ vô cùng cồng kềnh.\nCreative Synthesis: Quyết định chia nhỏ thành 3 notebook tách biệt: 01_attribution_basic.ipynb (cho RQ1), 02_logistic_benchmark.ipynb (cho RQ2), và 04_simulation.ipynb (cho RQ3), kết nối với nhau thông qua các file dữ liệu trung gian và dùng chung module cấu hình notebook_header.py.\nDecision Ownership: Chọn cấu trúc modular 3 notebook để tối ưu hóa sự minh bạch trong kiểm tra mã nguồn (reproducibility) và giảm thiểu rủi ro xung đột bộ nhớ khi huấn luyện mô hình.",
        "evidence": "Cấu trúc file trong thư mục notebooks gồm: 01_attribution_basic.ipynb, 02_logistic_benchmark.ipynb, 04_simulation.ipynb."
    },
    {
        "num": "4",
        "type": "DECISION",
        "stage": "Data Understanding & Preparation",
        "context": "Quyết định tích hợp dữ liệu trung gian từ đồng đội để đẩy nhanh tiến độ làm mô hình mà vẫn đảm bảo tính nhất quán của pipeline.",
        "prompt": "Bạn Khang cùng nhóm đã chuẩn bị xong 3 file CSV có sẵn các đặc trưng pre-computed cho First/Last touch và dummy variables. Nhưng thầy giáo khuyên nên viết code tự clean và tự sinh các biến này từ file raw để đảm bảo tính độc lập. Em nên dùng file của Khang hay tự viết lại?",
        "response": "AI khuyên em nên tự viết lại từ đầu bằng Pandas để chứng minh khả năng code cá nhân trong phần đánh giá và tránh phụ thuộc vào dữ liệu của thành viên khác.",
        "reflection": "Critical Thinking: AI đưa ra lời khuyên máy móc. Việc tự code lại toàn bộ các bước tính toán Linear_Weight hay one-hot encoder của Khang là sự lãng phí thời gian lớn (duplicate effort) khi dự án đang chạy đua với thời gian, và có rủi ro tạo ra sai lệch số liệu giữa các thành viên.\nContextualization: Nhân đóng vai trò quản lý chính và viết phần code phân tích chính. Khang đã viết script data_cleaning.py rất chi tiết và đã được cả nhóm thống nhất.\nCreative Synthesis: Quyết định sử dụng trực tiếp 3 file CSV nội bộ của Khang để làm đầu vào cho các notebook phân tích, đồng thời tạo file notebook_header.py để quản lý tập trung đường dẫn và logic load dữ liệu, đảm bảo tính nhất quán tuyệt đối.\nDecision Ownership: Đồng ý dùng kết quả trung gian của Khang làm nền tảng đầu vào để đẩy nhanh tiến độ làm mô hình, đảm bảo tính phân công công việc rõ ràng.",
        "evidence": "notebook_header.py chứa hàm load_touchpoints, load_journeys, load_encoded đọc trực tiếp từ data_touchpoints.csv, data_journeys.csv, data_encoded.csv."
    },
    {
        "num": "5",
        "type": "PROBLEM-SOLVING",
        "stage": "Data Understanding & Preparation",
        "context": "Phát hiện lỗi sắp xếp thứ tự touchpoint không đồng nhất do tệp tin CSV bị Excel tự động lược bỏ giây trong cột Timestamp.",
        "prompt": "Khi em chạy sort timestamp để xác định First/Last touch của các hành trình, em thấy thứ tự touchpoint bị xáo trộn nhẹ mỗi lần chạy và số liệu First-Touch bị lệch so với ban đầu. Tại sao lại như vậy?",
        "response": "AI dự đoán do thuật toán sắp xếp của Pandas (sort_values) mặc định là quicksort và không ổn định (unstable sort) đối với các giá trị trùng lặp. AI gợi ý chuyển sang dùng kind='mergesort'.",
        "reflection": "Critical Thinking: AI gợi ý đổi thuật toán sort nhưng không nhận ra nguyên nhân cốt lõi là dữ liệu đầu vào. Timestamp thực tế không trùng lặp ở mức giây, nhưng do Excel đã tự động strip mất phần giây khi Khang mở file thô để xem và lưu lại (e.g. \"07:58:51\" thành \"07:58\"), làm phát sinh hàng loạt timestamp trùng nhau ở mức phút.\nContextualization: Nhóm cần tính toán chính xác tuyệt đối First-Touch và Last-Touch để chạy attribution. Việc đổi sang mergesort chỉ che giấu lỗi chứ không sửa được lỗi mất dữ liệu giây.\nCreative Synthesis: Phát hiện ra hành vi tự động định dạng của Excel trên máy Khang. Giải pháp là tải lại file dữ liệu gốc từ nguồn Kaggle, cấm mở bằng Excel trực tiếp, viết code Python đọc trực tiếp và chạy lại toàn bộ script làm sạch của Khang.\nDecision Ownership: Quyết định hủy bỏ file CSV bị lỗi của Excel, chạy lại pipeline làm sạch trực tiếp từ file zip raw để giữ nguyên định dạng timestamp có giây.",
        "evidence": "notebook_header.py:line 34 tham số parse_dates=[\"Timestamp\"] và asserts sắp xếp tăng dần trong 01_attribution_basic.ipynb."
    },
    {
        "num": "6",
        "type": "PROBLEM-SOLVING",
        "stage": "Data Understanding & Preparation",
        "context": "Khắc phục lỗi đọc dữ liệu Boolean của Pandas khi các giá trị bool được lưu dưới dạng chuỗi văn bản chuỗi \"True\"/\"False\" trong file CSV.",
        "prompt": "Em đọc file CSV của Khang và filter cột Converted == True thì ra rỗng, nhưng dùng == 'True' thì lại ra kết quả. Làm sao để pandas tự động cast đúng kiểu bool?",
        "response": "AI hướng dẫn dùng pd.read_csv('file.csv', dtype={'Converted': bool}).",
        "reflection": "Critical Thinking: AI đưa ra hàm dtype nhưng thực tế Pandas sẽ cast mọi chuỗi không rỗng (kể cả chuỗi \"False\") thành True trong Python, dẫn đến toàn bộ hành trình đều bị coi là converted (leakage dữ liệu cực kỳ nghiêm trọng).\nContextualization: Khang ghi file CSV từ DataFrame bằng Python nhưng đã chuyển cột bool thành chuỗi do lỗi định dạng lưu trữ trung gian.\nCreative Synthesis: Viết hàm ánh xạ thủ công _BOOL_MAP = {'True': True, 'False': False} và định nghĩa hàm _coerce_bool dùng chung trong notebook_header.py để tự động sửa chữa kiểu dữ liệu cho toàn bộ 3 CSV đầu vào.\nDecision Ownership: Quyết định tự viết bộ lọc kiểu dữ liệu (coercion helper) trong module header chung thay vì sửa thủ công từng notebook để tránh lỗi lặp lại.",
        "evidence": "Mã nguồn notebook_header.py:line 19-28 chứa định nghĩa _BOOL_MAP và _coerce_bool()."
    },
    {
        "num": "7",
        "type": "DECISION",
        "stage": "Data Understanding & Preparation",
        "context": "Quyết định tối giản hóa mã nguồn làm sạch dữ liệu trung tâm bằng cách loại bỏ các đặc trưng không bắt buộc nhằm giảm thiểu xung đột tiến độ.",
        "prompt": "Em định bổ sung thêm các biến thời gian như hour, day_of_week và one-hot encoding cho Campaign vào file data_cleaning.py của Khang để dùng cho Regression. Việc này có cần thiết không?",
        "response": "AI khuyên nên tích hợp toàn bộ các đặc trưng phái sinh này vào script làm sạch trung tâm để dữ liệu luôn sẵn sàng cho mọi mô hình.",
        "reflection": "Critical Thinking: AI đưa ra lời khuyên mang tính lý thuyết hệ thống nhưng bỏ qua khía cạnh thực tế: timeline của nhóm rất gấp. Việc yêu cầu Khang sửa lại script trung tâm và chạy lại từ đầu sẽ mất thời gian phân phối lại công việc và kiểm tra chéo (regression testing).\nContextualization: Trong khi đó, các biến Campaign thực tế không được yêu cầu bởi 3 RQ chính, và các biến thời gian có thể sinh ra tức thời (on-the-fly) chỉ bằng 1 dòng lệnh Pandas trong notebook khi thực sự cần.\nCreative Synthesis: Quyết định viết hàm bổ sung tính năng thời gian cục bộ add_time_features trong notebook_header.py để chỉ gọi khi phân tích regression trong notebook 02, giữ nguyên script data_cleaning.py của Khang không đổi.\nDecision Ownership: Quyết định tối giản hóa dữ liệu làm sạch trung tâm để tránh xung đột timeline của nhóm.",
        "evidence": "Hàm add_time_features được viết tại notebook_header.py:line 59-70."
    },
    {
        "num": "8",
        "type": "DECISION",
        "stage": "Exploratory Data Analysis (EDA)",
        "context": "Thiết lập cơ chế kiểm soát chất lượng dữ liệu tự động ở đầu quy trình xử lý của mỗi notebook để ngăn chặn lỗi âm thầm khi cập nhật dữ liệu.",
        "prompt": "Làm sao để đảm bảo dữ liệu đầu vào của các notebook không bị thay đổi bất ngờ khi các thành viên khác cập nhật file CSV?",
        "response": "AI gợi ý viết mã in ra hình dạng DataFrame bằng print(df.shape) ở đầu notebook và kiểm tra trực quan bằng mắt thường mỗi lần chạy.",
        "reflection": "Critical Thinking: AI khuyên kiểm tra trực quan rất dễ bỏ sót khi chạy nhanh qua các cell (silent failure). Nếu số dòng hoặc tỷ lệ nhãn conversion bị thay đổi âm thầm do Khang sửa script làm sạch, các kết quả thống kê và regression sẽ bị sai lệch mà không có cảnh báo lỗi.\nContextualization: Nhân là người chịu trách nhiệm chính về tính chính xác của các kết quả phân tích trong paper. Cần một cơ chế tự động dừng luồng chạy nếu dữ liệu sai lệch.\nCreative Synthesis: Đưa các câu lệnh khẳng định nghiêm ngặt assert len(df_tp) == 10000 và assert df_jr['Converted'].sum() == 2381 lên cell đầu tiên của mọi notebook ngay sau khi load data.\nDecision Ownership: Quyết định thiết lập hệ thống cảnh báo tự động bằng Assertion thay vì chỉ in thông tin ra màn hình để bảo vệ tính nhất quán học thuật.",
        "evidence": "File 01_attribution_basic.ipynb cell load data chứa các câu lệnh assert."
    },
    {
        "num": "9",
        "type": "PROBLEM-SOLVING",
        "stage": "Exploratory Data Analysis (EDA)",
        "context": "Giải thích nghịch lý về tỷ lệ chuyển đổi (Conversion Rate) cao bất thường ở mức touchpoint (49.4%) so với thực tế các trang e-commerce.",
        "prompt": "Trong dataset Multi-touch Attribution của em, tỷ lệ Conversion = Yes ở mức dòng (touchpoint-level) lên tới 49.4%, và tỷ lệ user có ít nhất một Conversion = Yes lên tới 83.6%. Số liệu này có bình thường đối với một trang thương mại điện tử không? Em nên giải thích thế nào trong báo cáo?",
        "response": "AI khẳng định đây là tỷ lệ chuyển đổi mua hàng (purchase conversion rate) cực kỳ tốt của một doanh nghiệp thương mại điện tử hàng đầu và khuyên em nên dùng nó làm bằng chứng cho hiệu năng xuất sắc của hệ thống marketing.",
        "reflection": "Critical Thinking: Phát hiện AI đã hiểu sai ngữ cảnh kinh doanh (Context Misunderstanding). Trong thực tế thương mại điện tử, tỷ lệ mua hàng trung bình chỉ dao động từ 1% đến 3%. Tỷ lệ 83.6% user-level conversion là phi thực tế nếu định nghĩa \"Yes\" là \"mua hàng thành công\".\nContextualization: Dữ liệu này thực chất định nghĩa \"Yes\" là một tương tác tích cực tại touchpoint (ví dụ: click vào link, điền form, xem sản phẩm sâu) chứ không phải là giao dịch thanh toán cuối cùng.\nCreative Synthesis: Định nghĩa lại thuật ngữ trong bài paper: \"Touch-level Conversion (Yes)\" là \"Tương tác chuyển đổi tích cực tại điểm chạm\", còn \"Journey-level Converted User\" là \"Người dùng có phát sinh hành động tương tác sâu trong suốt chuỗi hành trình\".\nDecision Ownership: Quyết định làm rõ khái niệm này trong phần Data Understanding của paper để tránh việc giảng viên chất vấn về tính thực tế của số liệu kinh doanh.",
        "evidence": "DAP391_Analysis_Report.html phần 3.3 (Limitations) giải thích định nghĩa lại của nhãn conversion."
    },
    {
        "num": "10",
        "type": "VERIFICATION",
        "stage": "Exploratory Data Analysis (EDA)",
        "context": "Kiểm chứng khẳng định của AI về việc các mô hình Heuristics bắt buộc phải cho phân phối đóng góp khác biệt lớn khi độ dài hành trình tăng. (Liên kết với Hallucination Case 2)",
        "prompt": "AI nói với em rằng trên tập dữ liệu customer journey thực tế, mô hình First-Touch và Last-Touch chắc chắn sẽ cho ra kết quả attribution share khác nhau xa cho kênh Display Ads vì vị trí tác động khác nhau. Em có cần chạy Spearman correlation hay KL-divergence để so sánh không?",
        "response": "AI khẳng định chắc chắn có sự khác biệt rõ rệt (hệ số tương quan Spearman sẽ < 0.6) và khuyên nên trực quan hóa sự khác biệt lớn này bằng biểu đồ cột.",
        "reflection": "Critical Thinking: Phát hiện AI đã \"hallucinate\" về mức độ khác biệt của phân phối do AI mặc định dữ liệu có signal mạnh. Thực tế, khi chạy kiểm thử Spearman ρ giữa First và Last touch, hệ số tương quan đạt tới 0.98, và KL-divergence chỉ ở mức ~0.01 (gần như trùng khít hoàn toàn).\nContextualization: Điều này phản ánh đặc tính phân phối nhãn conversion của dữ liệu thô gần như đồng nhất và độc lập với vị trí chạm.\nCreative Synthesis: Chuyển đổi kết quả \"trùng khít\" này thành một luận điểm phân tích quan trọng trong paper: dữ liệu thô thiếu tín hiệu phân biệt, từ đó giải thích vì sao cần chuyển sang mô hình Logistic Regression và Markov Chain để khai thác tương tác sâu hơn.\nDecision Ownership: Quyết định phản bác khẳng định của AI, giữ nguyên kết quả thực nghiệm và đưa phân tích Spearman ρ/KL-divergence vào phần thảo luận hạn chế của nghiên cứu.",
        "evidence": "outputs/attribution_share.csv cho thấy sự phân bổ phần trăm giữa các model cực kỳ sát sao (FT: 17.3%, LT: 17.8%, Linear: 17.1% cho Direct Traffic)."
    },
    {
        "num": "11",
        "type": "DECISION",
        "stage": "Modeling & Regression Analysis",
        "context": "Lựa chọn thư viện hồi quy phù hợp trong Python để phục vụ mục tiêu giải thích ý nghĩa thống kê và kiểm định giả thuyết thay vì dự báo thuần túy.",
        "prompt": "Em muốn làm logistic regression để đo lường đóng góp của từng kênh marketing. Em nên dùng scikit-learn hay statsmodels trong Python?",
        "response": "AI khuyên nên dùng sklearn.linear_model.LogisticRegression vì nó chuẩn công nghiệp, dễ tích hợp vào pipeline machine learning và hỗ trợ cross-validation tốt.",
        "reflection": "Critical Thinking: AI thiên vị scikit-learn cho mục đích dự báo (prediction) nhưng bỏ qua yêu cầu cốt lõi của một nghiên cứu học thuật: giải thích ý nghĩa thống kê (inference). Scikit-learn mặc định thêm regularization (L2) làm méo hệ số coefficients, và không cung cấp trực tiếp p-value hay khoảng tin cậy (Confidence Interval) của các biến.\nContextualization: Nghiên cứu cần chứng minh p-value của các kênh marketing để chỉ ra kênh nào thực sự có tác động ý nghĩa đến conversion (RQ2).\nCreative Synthesis: Quyết định chọn thư viện statsmodels.api.Logit và sử dụng model.fit().summary() để trích xuất trực tiếp bảng hệ số, p-value, odds ratio và khoảng tin cậy 95% không bị chính quy hóa.\nDecision Ownership: Lựa chọn statsmodels làm công cụ hồi quy cốt lõi vì mục tiêu của dự án là kiểm định giả thuyết khoa học chứ không phải tối ưu hóa độ chính xác dự báo thuần túy.",
        "evidence": "File 02_logistic_benchmark.ipynb chứa mã nguồn gọi sm.Logit và xuất bảng hệ số ra outputs/logit_summary.csv."
    },
    {
        "num": "12",
        "type": "DECISION",
        "stage": "Modeling & Regression Analysis",
        "context": "Xác định cấp độ tổng hợp dữ liệu (Aggregation Level) phù hợp cho mô hình hồi quy để tránh rò rỉ thông tin (Data Leakage).",
        "prompt": "Trong dataset có cột Conversion của từng touchpoint và cột Converted của cả user. Khi chạy logistic regression để xem ảnh hưởng của kênh, em nên đưa 10,000 touchpoint vào trực tiếp hay gộp thành 2,847 user?",
        "response": "AI khuyên nên huấn luyện mô hình trực tiếp trên 10,000 dòng touchpoint vì kích thước dữ liệu lớn hơn sẽ giúp mô hình hồi quy học chính xác hơn và đạt AUC cao hơn.",
        "reflection": "Critical Thinking: AI đưa ra lời khuyên sai lầm nghiêm trọng về phương pháp luận. Biến mục tiêu Converted (Yes/No) thực chất là nhãn ở cấp độ người dùng (journey-level). Nếu sử dụng 10,000 touchpoint và lặp lại nhãn conversion của cùng một user trên nhiều touchpoint, mô hình sẽ gặp hiện tượng rò rỉ dữ liệu (data leakage) và vi phạm giả định độc lập của sai số hồi quy.\nContextualization: Dữ liệu Attribution bản chất là quan sát chuỗi hành vi của 2,847 khách hàng độc lập.\nCreative Synthesis: Gộp dữ liệu về mức người dùng bằng cách áp dụng phép toán logic .any() trên 6 cột dummy biến kênh (Channel_Social Media, v.v.), sau đó kết hợp với nhãn conversion duy nhất của journey để chạy hồi quy.\nDecision Ownership: Quyết định thực hiện hồi quy ở cấp độ user-journey (2,847 quan sát) để đảm bảo tính chặt chẽ về mặt thống kê của nghiên cứu.",
        "evidence": "Mã nguồn hồi quy trong 02_logistic_benchmark.ipynb gom nhóm theo user_id và lấy .max() cho các kênh dummies trước khi train."
    },
    {
        "num": "13",
        "type": "VERIFICATION",
        "stage": "Modeling & Regression Analysis",
        "context": "Đánh giá đề xuất của AI về việc áp dụng các kỹ thuật cân bằng dữ liệu mạnh (SMOTE/Class Weights) khi nhãn chuyển đổi bị lệch nhẹ (83.6% vs 16.4%). (Liên kết với Hallucination Case 3)",
        "prompt": "Tỷ lệ conversion trong tập dữ liệu của em là khoảng 83.6% converted và 16.4% non-converted. Em có cần xử lý mất cân bằng dữ liệu khi chia train/test không?",
        "response": "AI đề xuất dùng phương pháp SMOTE để sinh thêm dữ liệu cho lớp thiểu số hoặc dùng class_weight='balanced' để bù đắp sự mất cân bằng.",
        "reflection": "Critical Thinking: AI đã phản ứng quá đà trước hiện tượng mất cân bằng dữ liệu (imbalance). Tỷ lệ 83:17 là mức mất cân bằng cực kỳ nhẹ (mild imbalance), hoàn toàn không cần can thiệp bằng các kỹ thuật phức tạp như SMOTE vốn có thể làm sai lệch phân phối thực tế của hệ số hồi quy.\nContextualization: Mục tiêu là ước lượng đúng odds ratio thực tế để ra quyết định phân bổ ngân sách thực.\nCreative Synthesis: Quyết định không áp dụng SMOTE hay class weighting, chỉ sử dụng phân tách tập train/test theo tỷ lệ 70/30 có kèm tham số stratify=y trong hàm train_test_split của scikit-learn để đảm bảo tỷ lệ nhãn conversion đồng đều ở cả hai tập.\nDecision Ownership: Quyết định giữ nguyên phân phối nhãn thô và chỉ áp dụng stratify để đảm bảo tính khách quan của phép thử hồi quy.",
        "evidence": "02_logistic_benchmark.ipynb chứa dòng code train_test_split(..., test_size=0.3, stratify=y, random_state=42)."
    },
    {
        "num": "14",
        "type": "DECISION",
        "stage": "Modeling & Regression Analysis",
        "context": "Thiết lập hệ chỉ số đánh giá độ khớp của mô hình (Goodness-of-Fit) phù hợp cho tập nhãn mất cân bằng để tránh bẫy độ chính xác (Accuracy Paradox).",
        "prompt": "Làm sao để đánh giá mô hình Logistic Regression của em có giải thích tốt hành vi chuyển đổi của khách hàng hay không khi dữ liệu có p-value các kênh rất lớn?",
        "response": "AI khuyên chỉ nên sử dụng Accuracy (độ chính xác phân loại) làm metric chính vì giảng viên dễ hiểu và dễ báo cáo trong slide.",
        "reflection": "Critical Thinking: AI đề xuất dùng Accuracy là sai lầm đối với tập dữ liệu mất cân bằng nhẹ. Nếu mô hình đoán mò toàn bộ là \"Converted\" (đoán True), Accuracy vẫn đạt tới 83.6% nhưng mô hình hoàn toàn vô dụng.\nContextualization: Nghiên cứu học thuật đòi hỏi phải chứng minh được khả năng phân biệt (discriminative power) và độ khớp (goodness-of-fit).\nCreative Synthesis: Kết hợp song song hai chỉ số: McFadden's Pseudo-R² để đo tỷ lệ giải thích phương sai so với mô hình rỗng, và AUC (Area Under ROC Curve) trên tập test độc lập để đo năng lực phân loại thực tế của mô hình.\nDecision Ownership: Chọn bộ đôi Pseudo-R² và AUC làm thước đo chính thức để đánh giá chất lượng mô hình hồi quy, bỏ qua chỉ số Accuracy dễ gây hiểu lầm.",
        "evidence": "Kết quả lưu tại outputs/regression_metrics.csv với các giá trị McFadden's R2 và AUC."
    },
    {
        "num": "15",
        "type": "DECISION",
        "stage": "Modeling & Regression Analysis",
        "context": "Chọn thuật toán đo lường mức độ tương đồng giữa xếp hạng đóng góp kênh từ các mô hình Heuristics và hồi quy hồi quy.",
        "prompt": "Em muốn so sánh xem thứ tự xếp hạng độ quan trọng của 6 kênh giữa mô hình hồi quy và mô hình Linear Attribution có giống nhau không. Em nên dùng Pearson correlation hay Spearman?",
        "response": "AI khuyên nên dùng Pearson correlation vì đây là hệ số tương quan phổ biến nhất và đo lường trực tiếp mối quan hệ tuyến tính giữa hai chuỗi giá trị.",
        "reflection": "Critical Thinking: AI khuyên dùng Pearson là không phù hợp vì giá trị hệ số coefficients của Logistic Regression (odds ratios) và phần trăm đóng góp của Heuristic models nằm trên hai thang đo hoàn toàn khác nhau về biên độ và phân phối. Việc so sánh giá trị tuyến tính trực tiếp là vô nghĩa.\nContextualization: Nghiên cứu chỉ cần kiểm chứng liệu hai phương pháp có đồng thuận về \"thứ tự ưu tiên\" (ranking) của các kênh marketing để chia ngân sách hay không.\nCreative Synthesis: Áp dụng hệ số tương quan thứ hạng Spearman's rank correlation (scipy.stats.spearmanr) trên danh sách xếp hạng của 6 kênh từ 1 đến 6 để đo mức độ đồng thuận về mặt thứ tự.\nDecision Ownership: Chọn Spearman's ρ làm metric chính để kiểm chứng sự tương đồng trong định hướng phân bổ ngân sách của các mô hình.",
        "evidence": "File 02_logistic_benchmark.ipynb chứa cell tính toán scipy.stats.spearmanr và lưu kết quả tương quan."
    },
    {
        "num": "16",
        "type": "DECISION",
        "stage": "Evaluation, Visualization & Reporting",
        "context": "Giới hạn số lượng kịch bản mô phỏng ngân sách (Budget Scenarios) để duy trì tính mạch lạc và chiều sâu của bài viết.",
        "prompt": "Khi chạy mô hình giả lập ngân sách marketing (RQ3), em có nên đề xuất thêm kịch bản phân bổ theo trọng số First-Touch và Last-Touch để làm phong phú bài viết không?",
        "response": "AI khuyến khích nên viết thêm ít nhất 2 kịch bản nữa dựa trên First-Touch và Last-Touch để bài nghiên cứu trông đồ sộ, có nhiều biểu đồ so sánh hơn.",
        "reflection": "Critical Thinking: AI đề xuất tăng số lượng kịch bản để tăng dung lượng bài nhưng bỏ qua tính khả thi và yêu cầu ban đầu. Thêm quá nhiều kịch bản heuristic trung gian chỉ làm loãng thông điệp chính và khiến phần phân tích độ nhạy (sensitivity analysis) trở nên cực kỳ rối rắm mà không đem lại insight mới.\nContextualization: Bài paper định hướng theo khuôn mẫu đã thống nhất của nhóm và guideline của giảng viên. Cần tập trung tối đa làm sâu 3 kịch bản chính đã được phê duyệt trong Project Planning mới.\nCreative Synthesis: Tập trung tối ưu hóa 3 kịch bản cốt lõi: S0 (Equal Split), S1 (CR weighted), S2 (Linear share weighted), đồng thời phát triển biểu đồ waterfall 2-panel và bảng phân tích độ nhạy ±20% thật chi tiết cho 3 kịch bản này.\nDecision Ownership: Quyết định giới hạn ở 3 kịch bản cốt lõi để đảm bảo sự mạch lạc và chiều sâu cho phần thảo luận kinh doanh.",
        "evidence": "File 04_simulation.ipynb chứa logic giả lập cho đúng 3 kịch bản S0, S1, S2."
    },
    {
        "num": "17",
        "type": "VERIFICATION",
        "stage": "Evaluation, Visualization & Reporting",
        "context": "Xác minh tính xác thực của tài liệu tham khảo khoa học do AI cung cấp về mô hình Markov Chain trong Multi-touch Attribution. (Liên kết với Hallucination Case 1)",
        "prompt": "Em muốn trích dẫn nguồn gốc của mô hình attribution dựa trên Markov Chain. AI hãy gợi ý cho em bài báo khoa học chuẩn nhất để đưa vào Related Work.",
        "response": "AI đề xuất trích dẫn bài báo: \"Shao, X. and Li, L. (2011). Data-driven multi-touch attribution models. Proceedings of the 17th ACM SIGKDD international conference on Knowledge discovery and data mining, 375-383.\"",
        "reflection": "Critical Thinking: Phát hiện AI đã trích dẫn đúng tên bài báo và hội thảo nhưng bịa đặt (fabricate) tên tác giả hoặc năm xuất bản của bài Markov Attribution gốc. Bài báo thực tế của Shao và Li năm 2011 là giới thiệu mô hình bagging/ensemble và logistic regression cho MTA, còn bài báo áp dụng Markov Chain đầu tiên cho MTA là của Eva Anderl et al. (2016) đăng trên tạp chí European Journal of Operational Research.\nContextualization: Đây là lỗi học thuật cực kỳ nghiêm trọng có thể khiến bài paper bị đánh trượt ngay lập tức do đạo văn hoặc bịa đặt tài liệu.\nCreative Synthesis: Tiến hành tra cứu thủ công trên Google Scholar bằng DOI và tên bài báo để kiểm chứng, thay thế toàn bộ trích dẫn Markov sai lệch bằng nguồn của Eva Anderl et al. (2016) và hiệu chỉnh lại phần tổng quan y văn (Literature Review).\nDecision Ownership: Quyết định loại bỏ hoàn toàn tài liệu tham khảo do AI sinh ra khi chưa được kiểm chứng độc lập trên Google Scholar.",
        "evidence": "Báo cáo DAP391_Analysis_Report.html phần 3.3 liệt kê tài liệu tham khảo chuẩn xác của dự án."
    },
    {
        "num": "18",
        "type": "VERIFICATION",
        "stage": "Evaluation, Visualization & Reporting",
        "context": "Kiểm chứng tuyên bố của AI về hiệu năng lý tưởng của mô hình hồi quy (AUC > 0.7) trên tập đặc trưng Dummy của dữ liệu độc lập. (Liên kết với Hallucination Case 3)",
        "prompt": "Em chạy mô hình Logistic Regression trên tập dữ liệu đã one-hot channel dummies nhưng chỉ số AUC chỉ đạt 0.505, gần như tương đương đoán mò. Có phải em bị lỗi code ở đâu không?",
        "response": "AI khẳng định chắc chắn có lỗi trong phần tiền xử lý dữ liệu hoặc chia tập train/test, vì \"Logistic Regression trên dữ liệu thực tế luôn phải đạt AUC ít nhất 0.70 khi sử dụng đầy đủ 6 kênh dummy marketing\".",
        "reflection": "Critical Thinking: Phát hiện AI đã \"hallucinate\" về hiệu năng mô hình (Logic Error/Oversimplification). Khi hai biến độc lập (Channel) và biến phụ thuộc (Conversion) độc lập tuyến tính với nhau (đã được chứng minh qua kiểm định Chi-square p > 0.85), bất kỳ mô hình phân loại nào (kể cả hồi quy phức tạp) cũng chỉ có thể cho ra kết quả tương đương đoán ngẫu nhiên (AUC ~ 0.5). Không có lỗi code nào ở đây.\nContextualization: Dữ liệu dự án thực tế phản ánh hành vi chuyển đổi của người dùng gần như không bị chi phối đơn giản bởi kênh tiếp xúc trực tiếp.\nCreative Synthesis: Giữ vững kết quả AUC 0.505 trong paper, dùng nó làm bằng chứng khoa học đắt giá để bác bỏ giả thuyết cho rằng các mô hình heuristics đơn giản có thể tự động tối ưu hóa ngân sách mà không cần xét đến tương tác chuỗi hành trình phức tạp hơn.\nDecision Ownership: Quyết định không cố gắng \"tăng khống\" AUC bằng cách xáo trộn dữ liệu hay overfit mô hình, bảo vệ tính thực tế của kết quả thực nghiệm trước hội đồng.",
        "evidence": "outputs/regression_metrics.csv lưu trữ giá trị AUC thực nghiệm đạt 0.505."
    }
]

# Write to sheet 2 starting from Row 4
for idx, entry in enumerate(entries, start=4):
    ws2[f'A{idx}'] = entry["num"]
    ws2[f'B{idx}'] = entry["type"]
    ws2[f'C{idx}'] = entry["stage"]
    ws2[f'D{idx}'] = entry["context"]
    ws2[f'E{idx}'] = entry["prompt"]
    ws2[f'F{idx}'] = entry["response"]
    ws2[f'G{idx}'] = entry["reflection"]
    ws2[f'H{idx}'] = entry["evidence"]

# ==============================================================================
# SHEET 3: 3. Hallucination Detection
# ==============================================================================
print("Populating Sheet 3: Hallucination Detection...")
ws3 = wb['3. Hallucination Detection']

# Overwrite starting from Row 4
hallucinations = [
    {
        "num": "17",
        "type": "Fabrication",
        "claim": "Bài báo \"Shao & Li (2011)\" đăng tại hội thảo ACM SIGKDD là nguồn gốc chính thức giới thiệu mô hình Markov Chain cho bài toán Multi-touch Attribution.",
        "reality": "Bài báo của Shao & Li (2011) thực chất viết về mô hình hồi quy Logistic có Regularization và phương pháp Bagging. Bài báo đầu tiên áp dụng Markov Chain cho MTA là của Eva Anderl et al. (2016) đăng trên tạp chí European Journal of Operational Research.",
        "how": "Tra cứu độc lập tên bài báo, tác giả và nội dung tóm tắt trên Google Scholar và Scopus khi viết chương Related Work.",
        "action": "Thay thế toàn bộ trích dẫn sai trong bài paper bằng nguồn tài liệu chuẩn xác của Anderl et al. (2016) và viết lại nội dung tổng quan lý thuyết."
    },
    {
        "num": "18",
        "type": "Oversimplification",
        "claim": "Mô hình hồi quy Logistic sử dụng các biến dummy của 6 kênh marketing chắc chắn phải đạt chỉ số AUC tối thiểu 0.70 trên tập dữ liệu thực tế.",
        "reality": "Trên tập dữ liệu dự án thực tế, do biến độc lập (Channel) và biến phụ thuộc (Conversion) gần như độc lập với nhau (Chi-square p > 0.85), mô hình hồi quy tối ưu nhất cũng chỉ đạt AUC quanh mức 0.50 (tương đương phân loại ngẫu nhiên).",
        "how": "Thực thi mã nguồn huấn luyện mô hình trong file Jupyter 02_logistic_benchmark.ipynb, kiểm tra giá trị AUC thực tế của tập test độc lập.",
        "action": "Giữ nguyên chỉ số AUC thực tế (0.505), đưa vào mục Discussion & Limitations như một phát hiện khoa học chứng minh sự độc lập của dữ liệu thô."
    },
    {
        "num": "13",
        "type": "Context Misunderstanding",
        "claim": "Tập dữ liệu bị mất cân bằng nghiêm trọng (Imbalanced Data) do tỷ lệ converted đạt 83.6%, yêu cầu bắt buộc phải chạy SMOTE hoặc thiết lập trọng số class_weight='balanced'.",
        "reality": "Tỷ lệ 83.6% và 16.4% chỉ là mất cân bằng ở mức rất nhẹ (mild imbalance). Việc lạm dụng SMOTE hoặc cân bằng trọng số trong hồi quy Logistic sẽ làm méo mó ước lượng odds ratios và khoảng tin cậy của các hệ số.",
        "how": "Kiểm tra tỷ lệ phân bố thực tế của nhãn, đối chiếu với các tài liệu hướng dẫn kinh tế lượng (Econometrics) chuẩn về xử lý mất cân bằng nhẹ.",
        "action": "Bỏ qua toàn bộ các kỹ thuật SMOTE hay class weights của AI, chỉ áp dụng phân tách train/test có đối xứng nhãn stratify=y trong scikit-learn."
    }
]

for idx, h in enumerate(hallucinations, start=4):
    ws3[f'A{idx}'] = h["num"]
    ws3[f'B{idx}'] = h["type"]
    ws3[f'C{idx}'] = h["claim"]
    ws3[f'D{idx}'] = h["reality"]
    ws3[f'E{idx}'] = h["how"]
    ws3[f'F{idx}'] = h["action"]

# Clear any placeholder text in Row 5 Column A (e.g. '[Add more...]')
ws3['A5'] = "18"
ws3['A6'] = "13"

# ==============================================================================
# SHEET 4: 4. Self-Assessment Checklist
# ==============================================================================
print("Populating Sheet 4: Self-Assessment Checklist...")
ws4 = wb['4. Self-Assessment Checklist']

# Checkboxes for Section A (Rows 6 to 10, Column 3 / C)
for r in range(6, 11):
    ws4.cell(row=r, column=3).value = "☑"

# Checkboxes for Section B (Rows 14 to 18, Column 3 / C)
for r in range(14, 19):
    ws4.cell(row=r, column=3).value = "☑"

# Current values for Section B (Column 4 / D)
ws4['D14'] = "18 Core Prompts"
ws4['D15'] = "Đủ 5 components (Business 3, Data Prep 4, EDA 3, Modeling 5, Evaluation 3)"
ws4['D16'] = "3 Hallucination Cases"
ws4['D17'] = "18/18 entries đầy đủ 4 câu hỏi Reflection"
ws4['D18'] = "18/18 entries có evidence chi tiết"

# Oral Vivas Prep (Rows 28 to 30)
# Cell contents:
# Row 28: Entry #11, Row 29: Entry #12, Row 30: Entry #2
oral_vivas = [
    ("11", "Có, dùng statsmodels để lấy p-value không bị L2 regularization bias", "Có, khuyên dùng scikit-learn vì là chuẩn công nghiệp", "File 02_logistic_benchmark.ipynb và logit_summary.csv"),
    ("12", "Có, để tránh data leakage vì Converted là journey-level", "Có, khuyên train trực tiếp trên 10k touchpoint dòng", "02_logistic_benchmark.ipynb groupby user_id before fit"),
    ("2", "Có, vì Chi-square test p > 0.85 cho thấy các kênh độc lập, forecasting vô nghĩa", "Có, khuyên train Random Forest dự đoán doanh số", "DAP391_Analysis_Report.html phần 3.1 & 3.3")
]

for idx, (entry_num, q1, q2, q3) in enumerate(oral_vivas, start=28):
    ws4.cell(row=idx, column=1).value = entry_num
    ws4.cell(row=idx, column=2).value = q1
    ws4.cell(row=idx, column=3).value = q2
    ws4.cell(row=idx, column=4).value = q3

# Save workbook
print("Saving updated workbook...")
wb.save(TEMPLATE_PATH)
print("SUCCESSFULLY COMPLETED!")
